from flask import Flask, request, jsonify, send_file
from openpyxl import load_workbook
import openpyxl
import os
import base64
from io import BytesIO
from flask_cors import CORS

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

def load_workbook_from_data(file_data):
    """Load workbook from base64 encoded data"""
    try:
        # Decode base64 string to bytes
        file_bytes = base64.b64decode(file_data)
        # Create BytesIO object
        excel_file = BytesIO(file_bytes)
        # Load workbook
        return load_workbook(excel_file)
    except Exception as e:
        raise Exception(f"Error loading workbook: {str(e)}")

@app.route('/get_students', methods=['POST'])
def get_students():
    try:
        data = request.get_json()
        print("Received data:", data)  # Debug print
        
        # Handle web upload (base64 data)
        if 'file_data' in data:
            print("Processing file data")  # Debug print
            workbook = load_workbook_from_data(data['file_data'])
            sheet1 = workbook['Sheet1']
            student_count = sheet1.max_row - 4
            workbook.close()
            return jsonify({'number_of_students': student_count})
            
        # Handle file path
        elif 'file_path' in data:
            print("Processing file path:", data['file_path'])  # Debug print
            filepath = data['file_path']
            if not os.path.exists(filepath):
                return jsonify({'error': 'File not found'}), 404
                
            workbook = load_workbook(filepath)
            sheet1 = workbook['Sheet1']
            student_count = sheet1.max_row - 4
            workbook.close()
            return jsonify({'number_of_students': student_count})
            
        else:
            print("No file data or path provided")  # Debug print
            return jsonify({'error': 'No file data or path provided'}), 400
            
    except Exception as e:
        print("Error:", str(e))  # Debug print
        return jsonify({'error': str(e)}), 500

@app.route('/calculate', methods=['POST'])
def calculate():
    try:
        data = request.get_json()
        print("Received calculate data:", data)  # Debug print
        
        min_value1 = data.get('min_value1', 0)  # attainment_01
        max_value1 = data.get('max_value1', 0)  # attainment_11
        max_value2 = data.get('max_value2', 0)  # attainment_02
        
        # Handle file path
        if 'file_path' in data:
            filepath = data['file_path']
            if not os.path.exists(filepath):
                return jsonify({'error': 'File not found'}), 404
            workbook = load_workbook(filepath)
            sheet1 = workbook['Sheet1']
            sheet2 = workbook['Sheet2']
            last_row = sheet1.max_row
            No_of_Students = last_row - 4

            # Process the calculations
            process_columns(sheet1, sheet2, last_row, min_value1, max_value1, max_value2)
            save_avg_to_another_cell(sheet1, sheet2)
            save_ese(sheet1, sheet2)
            save_tw(sheet1, sheet2)
            save_external_avg(sheet1, sheet2)
            calculate_avg_and_save(sheet1, start_col=28, end_col=41, start_row=5, end_row=10, avg_row=12)
            co_table1(sheet1, sheet2)
            calculate_avg_and_save(sheet2, start_col=9, end_col=22, start_row=11, end_row=17, avg_row=18)
            co_table2(sheet1, sheet2)
            calculate_avg_and_save(sheet2, start_col=24, end_col=37, start_row=11, end_row=17, avg_row=18)

            # Save the workbook
            workbook.save(filepath)
            workbook.close()

            return jsonify({
                'number_of_students': No_of_Students,
                'message': 'Data Calculated Successfully'
            })
        else:
            return jsonify({'error': 'No file path provided'}), 400

    except Exception as e:
        print("Calculate error:", str(e))  # Debug print
        return jsonify({'error': str(e)}), 500

@app.route('/api/download', methods=['POST'])
def download():
    try:
        data = request.get_json()
        filepath = data.get('filepath')

        if not filepath:
            return jsonify({'error': 'Filepath is required'}), 400

        if not os.path.exists(filepath):
            return jsonify({'error': 'File not found'}), 404

        return send_file(
            filepath,
            as_attachment=True,
            download_name=os.path.basename(filepath)
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500

def calculate_attainment_level(sheet1, col, threshold, last_row, No_of_Students):
    count = 0
    for cell in sheet1.iter_cols(min_col=col, max_col=col, min_row=4, max_row=last_row):
        for col_cell in cell:
            val = col_cell.value
            if val is not None:
                try:
                    if float(val) >= threshold:
                        count += 1
                except (TypeError, ValueError):
                    pass
    attainment = int((count / No_of_Students) * 100)
    return count, attainment

def count_empty_or_null_cells(sheet1, col, last_row):
    empty_count = 0
    for cell in sheet1.iter_cols(min_col=col + 2, max_col=col + 2, min_row=1 + 3, max_row=last_row):
        for col_cell in cell:
            if col_cell.value is None or col_cell.value == 0:
                empty_count += 1
    return empty_count

def save_attainment_results(sheet2, col, count, attainment, sheet2_row, attainment_01, attainment_11, attainment_02):
    sheet2.cell(row=sheet2_row + 1, column=col, value=count)
    sheet2.cell(row=sheet2_row + 2, column=col, value=attainment)
    
    if attainment_11 > attainment >= attainment_01:
        sheet2.cell(row=sheet2_row + 3, column=col, value=1)
    elif attainment_02 > attainment >= attainment_11:
        sheet2.cell(row=sheet2_row + 3, column=col, value=2)
    elif attainment >= attainment_02:
        sheet2.cell(row=sheet2_row + 3, column=col, value=3)
    else:
        sheet2.cell(row=sheet2_row + 3, column=col, value=0)

def calculate_avg_of_columns_in_row(sheet2, columns, row):
    total = 0
    count = 0
    for col in columns:
        cell_value = sheet2.cell(row=row, column=col).value
        if isinstance(cell_value, (int, float)):
            total += cell_value
            count += 1
    if count == 0:
        return 0
    avg = total / count
    return avg

def calculate_external_avg(sheet2, rows):
    column_value = sheet2.cell(row=rows, column=3).value
    tw = sheet2.cell(row=rows, column=5).value
    ESE = sheet2.cell(row=rows, column=4).value
    
    if isinstance(tw, (int, float)) and isinstance(ESE, (int, float)) and isinstance(column_value, (int, float)):
        Avg_1 = (0.3 * (column_value + tw) / 2) + (0.7 * ESE)
        return Avg_1
    else:
        return None

def save_external_avg(sheet1, sheet2):
    cal1 = calculate_external_avg(sheet2, 10)
    cal2 = calculate_external_avg(sheet2, 11)
    cal3 = calculate_external_avg(sheet2, 12)
    cal4 = calculate_external_avg(sheet2, 13)
    cal5 = calculate_external_avg(sheet2, 14)
    cal6 = calculate_external_avg(sheet2, 15)

    if cal1 is not None:
        sheet2.cell(row=10, column=6, value=round(cal1, 2))
    if cal2 is not None:
        sheet2.cell(row=11, column=6, value=round(cal2, 2))
    if cal3 is not None:
        sheet2.cell(row=12, column=6, value=round(cal3, 2))
    if cal4 is not None:
        sheet2.cell(row=13, column=6, value=round(cal4, 5))
    if cal5 is not None:
        sheet2.cell(row=14, column=6, value=round(cal5, 5))
    if cal6 is not None:
        sheet2.cell(row=15, column=6, value=round(cal6, 5))

def save_avg_to_another_cell(sheet1, sheet2):
    row_to_calculate = 6
    Co1, Co2, Co3, Co4, Co5, Co6 = [], [], [], [], [], []
    
    for col in range(3, 24):
        cell_value = sheet1.cell(row=4, column=col).value
        if cell_value == 1:
            Co1.append(col)
        elif cell_value == 2:
            Co2.append(col)
        elif cell_value == 3:
            Co3.append(col)
        elif cell_value == 4:
            Co4.append(col)
        elif cell_value == 5:
            Co5.append(col)
        elif cell_value == 6:
            Co6.append(col)

    avg1 = calculate_avg_of_columns_in_row(sheet2, Co1, row_to_calculate)
    avg2 = calculate_avg_of_columns_in_row(sheet2, Co2, row_to_calculate)
    avg3 = calculate_avg_of_columns_in_row(sheet2, Co3, row_to_calculate)
    avg4 = calculate_avg_of_columns_in_row(sheet2, Co4, row_to_calculate)
    avg5 = calculate_avg_of_columns_in_row(sheet2, Co5, row_to_calculate)
    avg6 = calculate_avg_of_columns_in_row(sheet2, Co6, row_to_calculate)

    sheet2.cell(row=10, column=3, value=round(avg1, 2))
    sheet2.cell(row=11, column=3, value=round(avg2, 2))
    sheet2.cell(row=12, column=3, value=round(avg3, 2))
    sheet2.cell(row=13, column=3, value=round(avg4, 2))
    sheet2.cell(row=14, column=3, value=round(avg5, 2))
    sheet2.cell(row=15, column=3, value=round(avg6, 2))

def process_columns(sheet1, sheet2, last_row, attainment_01, attainment_11, attainment_02):
    sheet2_row = 3  # Start writing results to Sheet2 from this row
    sheet1_row = 3  # Start reading from this row in Sheet1
    No_of_Students = last_row - 4
    
    for col in range(3, 43):
        val = sheet1.cell(row=sheet1_row, column=col).value
        if val is not None:
            threshold = float(val) * 0.6
        else:
            threshold = 0
        sheet2.cell(row=sheet2_row, column=col, value=threshold)
        count, attainment = calculate_attainment_level(sheet1, col, threshold, last_row, No_of_Students)
        
        # Save attainment results with the values
        sheet2.cell(row=sheet2_row + 1, column=col, value=count)
        sheet2.cell(row=sheet2_row + 2, column=col, value=attainment)
        
        # Save the attainment level using the passed values
        if attainment_11 > attainment >= attainment_01:
            sheet2.cell(row=sheet2_row + 3, column=col, value=1)
        elif attainment_02 > attainment >= attainment_11:
            sheet2.cell(row=sheet2_row + 3, column=col, value=2)
        elif attainment >= attainment_02:
            sheet2.cell(row=sheet2_row + 3, column=col, value=3)
        else:
            sheet2.cell(row=sheet2_row + 3, column=col, value=0)
            
        empty_count = count_empty_or_null_cells(sheet1, col, last_row)
        sheet2.cell(row=sheet2_row + 4, column=col, value=empty_count)

def save_ese(sheet1, sheet2):
    row_to_calculate = 6
    Co1, Co2, Co3, Co4, Co5, Co6 = [], [], [], [], [], []
    
    for col in range(24, 40):
        cell_value = sheet1.cell(row=4, column=col).value
        if cell_value == 1:
            Co1.append(col)
        elif cell_value == 2:
            Co2.append(col)
        elif cell_value == 3:
            Co3.append(col)
        elif cell_value == 4:
            Co4.append(col)
        elif cell_value == 5:
            Co5.append(col)
        elif cell_value == 6:
            Co6.append(col)

    avg1 = calculate_avg_of_columns_in_row(sheet2, Co1, row_to_calculate)
    avg2 = calculate_avg_of_columns_in_row(sheet2, Co2, row_to_calculate)
    avg3 = calculate_avg_of_columns_in_row(sheet2, Co3, row_to_calculate)
    avg4 = calculate_avg_of_columns_in_row(sheet2, Co4, row_to_calculate)
    avg5 = calculate_avg_of_columns_in_row(sheet2, Co5, row_to_calculate)
    avg6 = calculate_avg_of_columns_in_row(sheet2, Co6, row_to_calculate)

    sheet2.cell(row=10, column=4, value=round(avg1, 2))
    sheet2.cell(row=11, column=4, value=round(avg2, 2))
    sheet2.cell(row=12, column=4, value=round(avg3, 2))
    sheet2.cell(row=13, column=4, value=round(avg4, 2))
    sheet2.cell(row=14, column=4, value=round(avg5, 2))
    sheet2.cell(row=15, column=4, value=round(avg6, 2))

def save_tw(sheet1, sheet2):
    row_to_calculate = 6
    Co1 = [40, 41, 42]
    avg1 = calculate_avg_of_columns_in_row(sheet2, Co1, row_to_calculate)
    avg1_float = round(avg1, 2)

    for row in range(10, 16):
        sheet2.cell(row=row, column=5, value=avg1_float)

def calculate_avg_and_save(sheet, start_col, end_col, start_row, end_row, avg_row):
    for col in range(start_col, end_col + 1):
        total = 0
        count = 0
        for row in range(start_row, end_row + 1):
            cell_value = sheet.cell(row=row, column=col).value
            if isinstance(cell_value, (int, float)):
                total += cell_value
                count += 1
        if count > 0:
            avg = total / count
            avg_round = round(avg, 2)
        else:
            avg_round = 0
        sheet.cell(row=avg_row, column=col, value=avg_round)

def co_table1(sheet1, sheet2):
    for col in range(45, 59):
        for row in range(4, 10):
            cell = sheet1.cell(row=row, column=col)
            cell_value = cell.value
            val = sheet2.cell(row=row + 6, column=6).value

            if isinstance(cell_value, (int, float)) and isinstance(val, (int, float)):
                if cell_value == 3:
                    new_value = val
                elif cell_value == 2:
                    new_value = float(val) * 0.66
                elif cell_value == 1:
                    new_value = float(val) * 0.33
                else:
                    new_value = 0
            else:
                new_value = ' '

            new_row = row + 7
            new_col = col - 36
            sheet2.cell(row=new_row, column=new_col, value=new_value)

def co_table2(sheet1, sheet2):
    for col in range(28, 42):
        for row in range(5, 11):
            cell = sheet1.cell(row=row - 1, column=col + 17)
            cell_value = cell.value
            val = sheet2.cell(row=row + 5, column=6).value

            if isinstance(cell_value, (int, float)) and isinstance(val, (int, float)):
                if cell_value in [3, 1, 2]:
                    new_value = val
                else:
                    new_value = ' '
            else:
                new_value = ' '

            new_row = row + 6
            new_col = col - 4
            sheet2.cell(row=new_row, column=new_col, value=new_value)

if __name__ == '__main__':
    app.run(debug=True, port=5000) 